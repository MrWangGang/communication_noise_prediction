import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import pickle
import openpyxl
from openpyxl.styles import PatternFill
import random
import os
import sys

from torch.nn import TransformerEncoderLayer, TransformerEncoder

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def set_seed(seed):
    """
    设置所有可能的随机种子以确保实验的可复现性
    """
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False
    print(f"✅ Random seed set to {seed}")

import torch
import torch.nn as nn

class CrossAttentionLayer(nn.Module):
    def __init__(self, query_dim, key_dim, value_dim, num_heads):
        super().__init__()
        self.attention = nn.MultiheadAttention(embed_dim=query_dim, num_heads=num_heads, batch_first=True)
        self.norm = nn.LayerNorm(query_dim)

    def forward(self, query, key, value):
        attention_output, _ = self.attention(query=query, key=key, value=value)
        output = self.norm(query + attention_output)
        return output

class LstmTransformer(nn.Module):
    def __init__(self, input_size, lstm_hidden_size=64, lstm_layers=1, dropout=0.1,
                 transformer_nhead=8, transformer_dim_feedforward=256, transformer_layers=1, output_size=None):
        super().__init__()

        self.transformer_embed_dim = transformer_dim_feedforward // 2
        self.transformer_input_projection = nn.Linear(input_size, self.transformer_embed_dim)

        self.lstm_output_dim = lstm_hidden_size * 2
        self.lstm = nn.LSTM(input_size, lstm_hidden_size, lstm_layers,
                            batch_first=True, bidirectional=True, dropout=dropout)

        self.transformer_encoder_layer = TransformerEncoderLayer(
            d_model=self.transformer_embed_dim,
            nhead=transformer_nhead,
            dim_feedforward=transformer_dim_feedforward,
            dropout=dropout,
            batch_first=True
        )
        self.transformer_encoder = TransformerEncoder(self.transformer_encoder_layer, num_layers=transformer_layers)


        self.cross_attention_t2l = CrossAttentionLayer(
            query_dim=self.transformer_embed_dim,
            key_dim=self.lstm_output_dim,
            value_dim=self.lstm_output_dim,
            num_heads=transformer_nhead
        )


        self.lstm_output_projection = nn.Linear(self.lstm_output_dim, self.transformer_embed_dim)
        self.cross_attention_l2t = CrossAttentionLayer(
            query_dim=self.transformer_embed_dim,
            key_dim=self.transformer_embed_dim,
            value_dim=self.transformer_embed_dim,
            num_heads=transformer_nhead
        )

        self.fc_out = nn.Linear(2 * self.transformer_embed_dim, output_size if output_size is not None else input_size)

    def forward(self, x):
        lstm_full_sequence_out, _ = self.lstm(x)
        transformer_input = self.transformer_input_projection(x)
        transformer_full_sequence_out = self.transformer_encoder(transformer_input)


        transformer_query = transformer_full_sequence_out[:, -1, :].unsqueeze(1)
        fused_t2l = self.cross_attention_t2l(
            query=transformer_query,
            key=lstm_full_sequence_out,
            value=lstm_full_sequence_out
        )
        fused_t2l = fused_t2l.squeeze(1)


        lstm_query = lstm_full_sequence_out[:, -1, :].unsqueeze(1)
        lstm_projected_query = self.lstm_output_projection(lstm_query)
        fused_l2t = self.cross_attention_l2t(
            query=lstm_projected_query,
            key=transformer_full_sequence_out,
            value=transformer_full_sequence_out
        )
        fused_l2t = fused_l2t.squeeze(1)

        # 3. 拼接两个融合结果
        combined_out = torch.cat((fused_t2l, fused_l2t), dim=1)

        # 最终预测
        out = self.fc_out(combined_out)
        return out

def main(start_index, end_index):
    """
    主预测函数
    """
    # 检查超参数
    if end_index - start_index <= 10:
        print("❌ 错误：`end_index - start_index` 必须大于 10。")
        return

    # 设置预测时的随机种子
    set_seed(42)

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(f"Using device: {device}")

    # 定义文件路径，请根据实际情况修改
    data_path = './datasets/complex_interference_dataset.csv'
    # 修改模型保存路径以匹配新的模型结构
    model_save_path = './model/lstm_transformer/best_model_lstm_transformer.pth'
    norm_params_path = './model/lstm_transformer/norm_params_lstm_transformer.pkl'
    # 修改输出文件名以反映新的模型
    output_xlsx_path = f'./prediction_lstm_transformer_{start_index}_to_{end_index}.xlsx'

    # 训练时的超参数，用于重建模型结构
    look_back = 10
    # LSTM 模型超参数
    lstm_hidden_size = 128
    lstm_layers = 3
    dropout = 0.3

    # Transformer 模型超参数
    transformer_nhead = 4
    transformer_dim_feedforward = 512
    transformer_layers = 1

    # ========================= 1. 加载标准化参数 =========================
    if not os.path.exists(norm_params_path):
        print(f"❌ 错误：未找到标准化参数文件 {norm_params_path}。请先运行训练脚本。")
        return
    with open(norm_params_path, 'rb') as f:
        norm_params = pickle.load(f)
    mean_per_feature = norm_params['mean']
    std_per_feature = norm_params['std']

    # ========================= 2. 加载模型结构和权重 =========================
    try:
        first_row_data = pd.read_csv(data_path, header=None, skiprows=1, nrows=1).values.astype(np.float32)
        input_size = first_row_data.shape[1]
        output_size = input_size
    except FileNotFoundError:
        print(f"❌ 错误：未找到数据集文件 {data_path}。")
        return

    # 实例化新的模型
    model = LstmTransformer(
        input_size=input_size,
        lstm_hidden_size=lstm_hidden_size,
        lstm_layers=lstm_layers,
        dropout=dropout,
        transformer_nhead=transformer_nhead,
        transformer_dim_feedforward=transformer_dim_feedforward,
        transformer_layers=transformer_layers,
        output_size=output_size
    ).to(device)

    if not os.path.exists(model_save_path):
        print(f"❌ 错误：未找到模型权重文件 {model_save_path}。请先运行训练脚本。")
        return
    model.load_state_dict(torch.load(model_save_path, map_location=device))
    model.eval()
    print("✅ 模型和标准化参数加载成功。")

    # ========================= 3. 准备预测数据 =========================
    # 读取预测所需的完整数据，包括look_back部分
    data_start_row = start_index - look_back
    data_end_row = end_index

    try:
        data_df = pd.read_csv(
            data_path,
            header=None,
            skiprows=data_start_row,
            nrows=(data_end_row - data_start_row + 1)
        )
        all_data_for_prediction = data_df.values.astype(np.float32)
        print(f"✅ 成功读取从第 {data_start_row} 行到第 {data_end_row} 行的数据，共 {len(all_data_for_prediction)} 条。")
    except FileNotFoundError:
        print(f"❌ 错误：未找到数据集文件 {data_path}。")
        return

    # 将数据进行标准化处理
    normalized_data = (all_data_for_prediction - mean_per_feature) / std_per_feature

    # 构造滑动窗口
    sequences = [normalized_data[i : i + look_back] for i in range(len(normalized_data) - look_back)]

    # 优化：先将列表转换为一个大的 NumPy 数组，再转换为 Tensor
    sequences_np = np.array(sequences, dtype=np.float32)
    inputs_tensor = torch.tensor(sequences_np).to(device)

    # ========================= 4. 进行预测 =========================
    with torch.no_grad():
        predicted_normalized = model(inputs_tensor).cpu().numpy()

    # 预测结果反标准化
    predicted_data = predicted_normalized * std_per_feature + mean_per_feature
    print("✅ 预测完成。")

    # ========================= 5. 写入 Excel (不含表头) =========================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prediction vs Actual"

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    current_row = 1
    for i in range(len(predicted_data)):
        # 实际值是从 all_data_for_prediction 数组中提取的，索引为 i + look_back
        actual_data_point = all_data_for_prediction[i + look_back]
        ws.append(actual_data_point.tolist())

        # 预测值是循环的第 i 个元素
        predicted_data_point = predicted_data[i]
        ws.append(predicted_data_point.tolist())

        # 给预测行上色
        for cell in ws[current_row + 1]:
            cell.fill = red_fill

        current_row += 2
        # 添加空行作为分隔
        ws.append([])
        current_row += 1

    print(f"✅ 预测结果已成功保存到 {output_xlsx_path}")
    wb.save(output_xlsx_path)

if __name__ == '__main__':
    # 您可以在这里设置要预测的起始和结束索引
    # 示例：预测从第 1000 行到第 1100 行的数据
    start_index_to_predict = 1000
    end_index_to_predict = 1100
    main(start_index_to_predict, end_index_to_predict)