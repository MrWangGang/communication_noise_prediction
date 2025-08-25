import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import pickle
import openpyxl
from openpyxl.styles import PatternFill
import random
import os
import sys

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def set_seed(seed):
    """
    设置所有可能的随机种子以确保实验的可复现性
    """
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False
    print(f"✅ Random seed set to {seed}")

class Lstm(nn.Module):
    def __init__(self, input_size, lstm_hidden_size=64, lstm_layers=1, dropout=0.1, output_size=None):
        super().__init__()
        self.lstm = nn.LSTM(input_size, lstm_hidden_size, lstm_layers,
                            batch_first=True, bidirectional=True, dropout=dropout)
        self.fc_out = nn.Linear(lstm_hidden_size * 2, output_size if output_size is not None else input_size)

    def forward(self, x):
        lstm_out, _ = self.lstm(x)
        out = lstm_out[:, -1, :]
        out = self.fc_out(out)
        return out


def main(start_index, end_index):
    """
    主预测函数
    """
    # 检查超参数
    if end_index - start_index <= 10:
        print("❌ 错误：`end_index - start_index` 必须大于 10。")
        return

    # 设置预测时的随机种子
    set_seed(42)

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(f"Using device: {device}")

    # 定义文件路径，请根据实际情况修改
    data_path = './datasets/complex_interference_dataset.csv'
    model_save_path = './model/lstm/best_model_lstm.pth'
    norm_params_path = './model/lstm/norm_params_lstm.pkl'
    output_xlsx_path = f'./prediction_lstm_{start_index}_to_{end_index}.xlsx'

    # 模型超参数
    look_back = 10
    lstm_hidden_size = 128
    lstm_layers = 3
    dropout = 0.3

    # ========================= 1. 加载标准化参数 =========================
    if not os.path.exists(norm_params_path):
        print(f"❌ 错误：未找到标准化参数文件 {norm_params_path}。请先运行训练脚本。")
        return
    with open(norm_params_path, 'rb') as f:
        norm_params = pickle.load(f)
    mean_per_feature = norm_params['mean']
    std_per_feature = norm_params['std']

    # ========================= 2. 加载模型结构和权重 =========================
    try:
        first_row_data = pd.read_csv(data_path, header=None, skiprows=1, nrows=1).values.astype(np.float32)
        input_size = first_row_data.shape[1]
        output_size = input_size
    except FileNotFoundError:
        print(f"❌ 错误：未找到数据集文件 {data_path}。")
        return

    model = Lstm(
        input_size=input_size,
        lstm_hidden_size=lstm_hidden_size,
        lstm_layers=lstm_layers,
        dropout=dropout,
        output_size=output_size
    ).to(device)

    if not os.path.exists(model_save_path):
        print(f"❌ 错误：未找到模型权重文件 {model_save_path}。请先运行训练脚本。")
        return
    model.load_state_dict(torch.load(model_save_path, map_location=device))
    model.eval()
    print("✅ 模型和标准化参数加载成功。")

    # ========================= 3. 准备预测数据 =========================
    data_start_row = start_index - look_back
    data_end_row = end_index

    try:
        data_df = pd.read_csv(
            data_path,
            header=None,
            skiprows=data_start_row,
            nrows=(data_end_row - data_start_row + 1)
        )
        all_data_for_prediction = data_df.values.astype(np.float32)
        print(f"✅ 成功读取从第 {data_start_row} 行到第 {data_end_row} 行的数据，共 {len(all_data_for_prediction)} 条。")
    except FileNotFoundError:
        print(f"❌ 错误：未找到数据集文件 {data_path}。")
        return

    normalized_data = (all_data_for_prediction - mean_per_feature) / std_per_feature
    sequences = [normalized_data[i : i + look_back] for i in range(len(normalized_data) - look_back)]
    sequences_np = np.array(sequences, dtype=np.float32)
    inputs_tensor = torch.tensor(sequences_np).to(device)

    # ========================= 4. 进行预测 =========================
    with torch.no_grad():
        predicted_normalized = model(inputs_tensor).cpu().numpy()

    predicted_data = predicted_normalized * std_per_feature + mean_per_feature
    print("✅ 预测完成。")

    # ========================= 5. 写入 Excel =========================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prediction vs Actual"

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    current_row = 1
    for i in range(len(predicted_data)):
        actual_data_point = all_data_for_prediction[i + look_back]
        ws.append(actual_data_point.tolist())

        predicted_data_point = predicted_data[i]
        ws.append(predicted_data_point.tolist())

        for cell in ws[current_row + 1]:
            cell.fill = red_fill

        current_row += 2
        ws.append([])
        current_row += 1

    print(f"✅ 预测结果已成功保存到 {output_xlsx_path}")
    wb.save(output_xlsx_path)

if __name__ == '__main__':
    start_index_to_predict = 1000
    end_index_to_predict = 1100
    main(start_index_to_predict, end_index_to_predict)