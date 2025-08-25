import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import pickle
import openpyxl
from openpyxl.styles import PatternFill
import random
import os
import sys

from torch.nn import TransformerEncoderLayer, TransformerEncoder

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def set_seed(seed):
    """
    设置所有可能的随机种子以确保实验的可复现性
    """
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False
    print(f"✅ Random seed set to {seed}")


class CrossAttention(nn.Module):
    def __init__(self, embed_dim, num_heads, dropout=0.1):
        super().__init__()
        self.mha = nn.MultiheadAttention(embed_dim, num_heads, dropout, batch_first=True)
        self.norm = nn.LayerNorm(embed_dim)
        self.dropout = nn.Dropout(dropout)

    def forward(self, query, key, value):
        attn_output, _ = self.mha(query, key, value)
        x = self.norm(query + self.dropout(attn_output))
        return x

class lstmTransformer(nn.Module):
    def __init__(self, input_size, hidden_size=128, num_layers=2, dropout=0.3, num_heads=4, output_size=None):
        super().__init__()

        self.bilstm1 = nn.LSTM(input_size, hidden_size, num_layers, batch_first=True, dropout=dropout, bidirectional=True)
        self.bilstm2 = nn.LSTM(input_size, hidden_size, num_layers, batch_first=True, dropout=dropout, bidirectional=True)
        self.bilstm3 = nn.LSTM(input_size, hidden_size, num_layers, batch_first=True, dropout=dropout, bidirectional=True)

        embed_dim = hidden_size * 2
        # 定义三个独立的交叉注意力模块
        self.cross_attn1 = CrossAttention(embed_dim=embed_dim, num_heads=num_heads, dropout=dropout)
        self.cross_attn2 = CrossAttention(embed_dim=embed_dim, num_heads=num_heads, dropout=dropout)
        self.cross_attn3 = CrossAttention(embed_dim=embed_dim, num_heads=num_heads, dropout=dropout)

        final_output_size = output_size if output_size is not None else input_size
        # 融合后的维度是 embed_dim * 3
        self.fc_out = nn.Linear(embed_dim * 3, final_output_size)

    def forward(self, x):
        lstm_out1, _ = self.bilstm1(x)
        lstm_out2, _ = self.bilstm2(x)
        lstm_out3, _ = self.bilstm3(x)

        # 拼接所有BiLSTM的输出，作为Key和Value
        all_lstm_outputs = torch.cat([lstm_out1, lstm_out2, lstm_out3], dim=1)

        # 每个BiLSTM的输出都作为自己的Query
        fused_output1 = self.cross_attn1(query=lstm_out1, key=all_lstm_outputs, value=all_lstm_outputs)
        fused_output2 = self.cross_attn2(query=lstm_out2, key=all_lstm_outputs, value=all_lstm_outputs)
        fused_output3 = self.cross_attn3(query=lstm_out3, key=all_lstm_outputs, value=all_lstm_outputs)

        # 拼接三个融合后序列的最后一个时间步的输出
        final_output = torch.cat([
            fused_output1[:, -1, :],
            fused_output2[:, -1, :],
            fused_output3[:, -1, :]
        ], dim=1)

        out = self.fc_out(final_output)
        return out

def main(start_index, end_index):
    """
    主预测函数
    """
    # 检查超参数
    if end_index - start_index <= 10:
        print("❌ 错误：`end_index - start_index` 必须大于 10。")
        return

    # 设置预测时的随机种子
    set_seed(42)

    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(f"Using device: {device}")

    # 定义文件路径，请根据实际情况修改
    data_path = './datasets/complex_interference_dataset.csv'
    # 修改模型保存路径以匹配新的模型结构
    model_save_path = './model/lstm_transformer/best_model_lstm_parallel.pth'
    norm_params_path = './model/lstm_transformer/norm_params_lstm_parallel.pkl'
    # 修改输出文件名以反映新的模型
    output_xlsx_path = f'./prediction_lstm_parallel_{start_index}_to_{end_index}.xlsx'

    # 训练时的超参数，用于重建模型结构
    look_back = 10
    # 模型参数
    lstm_hidden_size = 128
    lstm_num_layers = 2
    dropout = 0.3
    num_heads = 4


    # ========================= 1. 加载标准化参数 =========================
    if not os.path.exists(norm_params_path):
        print(f"❌ 错误：未找到标准化参数文件 {norm_params_path}。请先运行训练脚本。")
        return
    with open(norm_params_path, 'rb') as f:
        norm_params = pickle.load(f)
    mean_per_feature = norm_params['mean']
    std_per_feature = norm_params['std']

    # ========================= 2. 加载模型结构和权重 =========================
    try:
        first_row_data = pd.read_csv(data_path, header=None, skiprows=1, nrows=1).values.astype(np.float32)
        input_size = first_row_data.shape[1]
        output_size = input_size
    except FileNotFoundError:
        print(f"❌ 错误：未找到数据集文件 {data_path}。")
        return

    # 实例化新的模型
    model = lstmTransformer(
        input_size=input_size,
        hidden_size=lstm_hidden_size,
        num_layers=lstm_num_layers,
        dropout=dropout,
        num_heads=num_heads,
        output_size=output_size
    ).to(device)

    if not os.path.exists(model_save_path):
        print(f"❌ 错误：未找到模型权重文件 {model_save_path}。请先运行训练脚本。")
        return
    model.load_state_dict(torch.load(model_save_path, map_location=device))
    model.eval()
    print("✅ 模型和标准化参数加载成功。")

    # ========================= 3. 准备预测数据 =========================
    # 读取预测所需的完整数据，包括look_back部分
    data_start_row = start_index - look_back
    data_end_row = end_index

    try:
        data_df = pd.read_csv(
            data_path,
            header=None,
            skiprows=data_start_row,
            nrows=(data_end_row - data_start_row + 1)
        )
        all_data_for_prediction = data_df.values.astype(np.float32)
        print(f"✅ 成功读取从第 {data_start_row} 行到第 {data_end_row} 行的数据，共 {len(all_data_for_prediction)} 条。")
    except FileNotFoundError:
        print(f"❌ 错误：未找到数据集文件 {data_path}。")
        return

    # 将数据进行标准化处理
    normalized_data = (all_data_for_prediction - mean_per_feature) / std_per_feature

    # 构造滑动窗口
    sequences = [normalized_data[i : i + look_back] for i in range(len(normalized_data) - look_back)]

    # 优化：先将列表转换为一个大的 NumPy 数组，再转换为 Tensor
    sequences_np = np.array(sequences, dtype=np.float32)
    inputs_tensor = torch.tensor(sequences_np).to(device)

    # ========================= 4. 进行预测 =========================
    with torch.no_grad():
        predicted_normalized = model(inputs_tensor).cpu().numpy()

    # 预测结果反标准化
    predicted_data = predicted_normalized * std_per_feature + mean_per_feature
    print("✅ 预测完成。")

    # ========================= 5. 写入 Excel (不含表头) =========================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prediction vs Actual"

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    current_row = 1
    for i in range(len(predicted_data)):
        # 实际值是从 all_data_for_prediction 数组中提取的，索引为 i + look_back
        actual_data_point = all_data_for_prediction[i + look_back]
        ws.append(actual_data_point.tolist())

        # 预测值是循环的第 i 个元素
        predicted_data_point = predicted_data[i]
        ws.append(predicted_data_point.tolist())

        # 给预测行上色
        for cell in ws[current_row + 1]:
            cell.fill = red_fill

        current_row += 2
        # 添加空行作为分隔
        ws.append([])
        current_row += 1

    print(f"✅ 预测结果已成功保存到 {output_xlsx_path}")
    wb.save(output_xlsx_path)

if __name__ == '__main__':
    # 您可以在这里设置要预测的起始和结束索引
    # 示例：预测从第 1000 行到第 1100 行的数据
    start_index_to_predict = 1000
    end_index_to_predict = 1100
    main(start_index_to_predict, end_index_to_predict)